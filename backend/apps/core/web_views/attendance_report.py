"""تقرير الحضور اليومي — تجميع بصمات الدخول والخروج."""
from datetime import datetime

from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET

from apps.attendance.models import AttendancePunch
from apps.attendance.selectors.daily_report import (
    build_daily_attendance_rows,
    daily_rows_to_table,
    summarize_daily_rows,
)
from apps.attendance.selectors.punch_export import punches_to_table_rows
from apps.attendance.selectors.punch_records import PUNCH_LIST_ORDERING, get_punch_queryset, get_punch_stats
from apps.attendance.selectors.biometric_devices import (
    filter_biometric_devices_for_user,
    get_biometric_devices_queryset,
)
from apps.core.decorators import permission_required
from apps.attendance.sub_permissions import ATTENDANCE_SCREEN_REPORT_VIEW
from apps.core.models import Branch
from apps.core.web_views._helpers import _user_accessible_branch_ids
from apps.core.web_views.attendance_records import (
    _apply_default_date_filters,
    _filters_to_querystring,
    _parse_filters,
)
from apps.employees.models import Employee


def _punches_for_report(request, filters: dict):
    from apps.attendance.selectors.employee_enrollment import (
        apply_employee_enrollment_to_filters,
        enrollment_filter_q,
    )

    date_from = None
    date_to = None
    if filters['date_from']:
        date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
    if filters['date_to']:
        date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()

    employee_id = filters['employee_id']
    employee_enrollments = []
    if employee_id:
        filters = apply_employee_enrollment_to_filters(filters, employee_id)
        employee_enrollments = filters.get('enrollments') or []
        employee_id = None

    qs = get_punch_queryset(
        device_id=filters['device_id'],
        branch_ids=filters['branch_ids'],
        employee_id=employee_id,
        device_user_id=filters['device_user_id'],
        date_from=date_from,
        date_to=date_to,
        punch_type=filters['punch_type'],
        mapped_only=filters['mapped_only'],
        search=filters['search'] or None,
    )
    qs = qs.filter(device_id__in=filter_biometric_devices_for_user(request.user).values('pk'))
    if employee_enrollments:
        qs = qs.filter(enrollment_filter_q(employee_enrollments))
    return qs


@permission_required(ATTENDANCE_SCREEN_REPORT_VIEW)
def attendance_report(request):
    from apps.core.utils.attendance_filters import clamp_attendance_date_range

    filters = _apply_default_date_filters(_parse_filters(request))
    filters, date_clamped = clamp_attendance_date_range(filters)
    if date_clamped:
        messages.warning(
            request,
            'تم تقييد فترة التقرير إلى 93 يوماً كحد أقصى لحماية الأداء.',
        )
    qs = _punches_for_report(request, filters)
    punch_stats = get_punch_stats(qs)

    load_daily = not request.GET.get('punches_page')
    all_rows: list = []
    from_daily_cache = False
    rows_truncated = False
    summary = summarize_daily_rows([], punch_total=punch_stats['total'])

    if load_daily:
        from apps.core.services.report_cache import (
            cache_bypass_requested,
            get_or_build_daily_attendance_rows,
        )

        bypass_cache = cache_bypass_requested(request)
        all_rows, from_daily_cache = get_or_build_daily_attendance_rows(
            user_id=request.user.id,
            filters=filters,
            bypass=bypass_cache,
            builder=lambda: build_daily_attendance_rows(qs),
        )
        rows_truncated = len(all_rows) >= 15_000
        summary = summarize_daily_rows(all_rows, punch_total=punch_stats['total'])
        if rows_truncated:
            messages.info(
                request,
                'تم عرض أول 15000 صف يومي فقط — ضيّق الفترة أو الفلاتر لعرض كامل.',
            )

    from apps.core.utils.pagination import clamp_page_size

    daily_per_page = clamp_page_size(request.GET.get('per_page'), default=50, maximum=200)
    daily_paginator = Paginator(all_rows, per_page=daily_per_page)
    daily_page = daily_paginator.get_page(request.GET.get('page') if load_daily else 1)

    punches_qs = qs.order_by(*PUNCH_LIST_ORDERING)
    punches_per_page = clamp_page_size(
        request.GET.get('punches_per_page'), default=100, maximum=200,
    )
    punches_paginator = Paginator(punches_qs, per_page=punches_per_page)
    punches_page = punches_paginator.get_page(request.GET.get('punches_page'))

    branches_qs = Branch.objects.filter(is_deleted=False, is_active=True).order_by('name')
    branch_ids = _user_accessible_branch_ids(request.user)
    if branch_ids is not None:
        branches_qs = branches_qs.filter(pk__in=branch_ids)

    mapped_filter = 'all'
    if filters['mapped_only'] is True:
        mapped_filter = 'yes'
    elif filters['mapped_only'] is False:
        mapped_filter = 'no'

    filter_employee = None
    if filters['employee_id']:
        filter_employee = Employee.objects.filter(
            pk=filters['employee_id'],
            is_deleted=False,
        ).select_related('branch', 'department').first()

    return render(request, 'pages/attendance/report.html', {
        'daily_page': daily_page,
        'punches_page': punches_page,
        'summary': summary,
        'punch_stats': punch_stats,
        'total_daily_rows': len(all_rows),
        'devices': get_biometric_devices_queryset(request.user),
        'branches': branches_qs,
        'employee_search_url': reverse('web:employee_picker_search'),
        'filter_employee': filter_employee,
        'filters': filters,
        'mapped_filter': mapped_filter,
        'querystring': _filters_to_querystring(filters),
        'daily_per_page': daily_per_page,
        'punches_per_page': punches_per_page,
        'punch_types': AttendancePunch.PunchType.choices,
        'date_range_clamped': date_clamped,
        'daily_rows_truncated': rows_truncated,
        'daily_from_cache': from_daily_cache,
    })


@permission_required(ATTENDANCE_SCREEN_REPORT_VIEW)
@require_GET
def attendance_report_export(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة.')
        return redirect('web:attendance_report')

    from apps.core.utils.attendance_filters import clamp_attendance_date_range

    filters = _apply_default_date_filters(_parse_filters(request))
    filters, _ = clamp_attendance_date_range(filters)
    qs = _punches_for_report(request, filters)
    daily_rows = build_daily_attendance_rows(qs)
    if len(daily_rows) > 5000:
        daily_rows = daily_rows[:5000]
    daily_table = daily_rows_to_table(daily_rows)
    punch_table = punches_to_table_rows(qs.order_by(*PUNCH_LIST_ORDERING))

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='1E40AF')

    from apps.attendance.selectors.punch_export import write_punch_table_sheet

    def _write_daily_sheet(ws, table: dict, title: str) -> None:
        ws.title = title
        ws.sheet_view.rightToLeft = True
        for col, h in enumerate(table['columns'], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = Font(bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='center')
        for row_idx, row in enumerate(table['rows'], 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col in range(1, len(table['columns']) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    ws_daily = wb.active
    _write_daily_sheet(ws_daily, daily_table, 'يومي')
    ws_detail = wb.create_sheet('تفصيلي')
    write_punch_table_sheet(ws_detail, punch_table)

    stamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'attendance_report_{stamp}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
