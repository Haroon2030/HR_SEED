"""تصدير بيانات الموظف إلى Excel ملوّن (بيانات تبويب الموظف فقط)."""
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from apps.employees.models import Employee
from apps.core.web_views._helpers import employee_branch_access_required
from apps.core.decorators import permission_required
from apps.core.salary_access import user_can_view_salary


def _employee_administration_label(employee):
    adm = getattr(employee, 'administration', None)
    if not adm:
        return None
    code = (getattr(adm, 'code', None) or '').strip()
    name = (getattr(adm, 'name', None) or '').strip()
    if code and name:
        return f'{code} — {name}'
    return code or name or None


@login_required
@permission_required('employees.view')
@employee_branch_access_required
def export_employee_salary_excel(request, employee_id):
    """يُنشئ ملف Excel ملوّن يحتوي على بيانات تبويب الموظف فقط."""
    from django.contrib import messages
    from django.shortcuts import redirect

    if not user_can_view_salary(request.user):
        messages.error(request, 'لا تملك صلاحية تصدير بيانات الرواتب.')
        return redirect('web:view_employee', employee_id=employee_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    employee = get_object_or_404(
        Employee.objects.select_related('administration'),
        id=employee_id,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الموظف"
    ws.sheet_view.rightToLeft = True

    # ──────────── أنماط ────────────
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='1E40AF')

    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')

    label_font = Font(name='Arial', size=11, bold=True, color='1E293B')
    label_fill = PatternFill('solid', fgColor='F1F5F9')

    value_font = Font(name='Arial', size=11, color='0F172A')
    value_fill = PatternFill('solid', fgColor='FFFFFF')

    salary_label_fill = PatternFill('solid', fgColor='DBEAFE')
    salary_value_fill = PatternFill('solid', fgColor='EFF6FF')

    allowance_fill = PatternFill('solid', fgColor='ECFDF5')
    allowance_label_fill = PatternFill('solid', fgColor='D1FAE5')

    deduction_fill = PatternFill('solid', fgColor='FEF2F2')
    deduction_label_fill = PatternFill('solid', fgColor='FEE2E2')

    total_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    total_fill = PatternFill('solid', fgColor='059669')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ──────────── بيانات الموظف كجدول أفقي (الأعمدة = الحقول) ────────────
    columns = [
        ("الاسم", employee.name),
        ("رقم الموظف", employee.employee_number),
        ("رقم الهوية", employee.id_number),
        ("الجوال", employee.phone),
        ("البريد الإلكتروني", employee.email),
        ("الجنسية", getattr(employee.nationality, 'name', None)),
        ("المهنة", getattr(employee.profession, 'name', None)),
        ("الكفالة", getattr(employee.sponsorship, 'name', None)),
        ("الفرع", getattr(employee.branch, 'name', None)),
        ("القسم", getattr(employee.department, 'name', None)),
        ("الإدارة", _employee_administration_label(employee)),
        ("مركز التكلفة", getattr(employee.cost_center, 'name', None)),
        ("تاريخ المباشرة", employee.hire_date),
        ("تاريخ التوقف", employee.end_date),
        ("تاريخ انتهاء التأمين الطبي", employee.medical_insurance_expiry_date),
        ("تاريخ انتهاء العقد (مخطط)", employee.contract_expiry_date),
        ("التأمين", getattr(employee.insurance, 'name', None)),
        ("فئة التأمين", getattr(employee.insurance_class, 'name', None)),
        ("الحالة", employee.get_status_display()),
        ("سبب الانتهاء", employee.end_reason),
    ]

    # عرض الأعمدة (مضغوط)
    for idx, (label, _v) in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 10

    n = len(columns)
    last_col = get_column_letter(n)

    # العنوان الكبير
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value="بيانات الموظف")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    c.border = border
    ws.row_dimensions[1].height = 38

    # تاريخ الإصدار
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=f"تاريخ الإصدار: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(name='Arial', size=10, italic=True, color='64748B')
    c.alignment = center
    ws.row_dimensions[2].height = 22

    # صف العناوين (الأعمدة) - بخط أصغر والتفاف للأسطر
    header_row = 4
    compact_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    for idx, (label, _v) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=idx, value=label)
        c.font = compact_header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 60

    # صف القيم
    value_row = header_row + 1
    compact_value_font = Font(name='Arial', size=9, color='0F172A')
    for idx, (_label, value) in enumerate(columns, start=1):
        c = ws.cell(row=value_row, column=idx,
                    value=value if value not in (None, '') else '—')
        c.font = compact_value_font
        c.fill = value_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[value_row].height = 22

    # تجميد صف العناوين
    ws.freeze_panes = ws.cell(row=value_row, column=1)

    # ──────────── إخراج الملف ────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = (employee.name or 'employee').replace(' ', '_')
    filename = f"employee_{safe_name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
