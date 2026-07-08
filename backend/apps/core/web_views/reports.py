"""
التقارير — بيانات تفصيلية بصفوف وأعمدة
كل تقرير يُرجع: columns (أعمدة) + rows (صفوف) + title
"""
from datetime import date, datetime, timedelta
from decimal import Decimal
from urllib.parse import urlencode

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q, F
from django.db.models.functions import Coalesce
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

from django.contrib import messages

from apps.core.decorators import any_permission_required, permission_required
from apps.core.permission_policy import report_allowed_for_user, user_can_view_financial_reports
from apps.core.models import Branch
from apps.core.filter_utils import apply_branch_filter, append_multi_param, parse_multi_filter_ids
from apps.core.web_views._helpers import _user_accessible_branch_ids
from apps.setup.models import Sponsorship
from apps.core.services.reports_catalog import PRIMARY_REPORT_SPECS, merge_reports_catalog

REPORT_GROUPS = [
    {'key': 'primary',      'title': 'التقارير',          'icon': 'bar-chart-3',   'color': 'primary',  'description': 'التقارير الأساسية المطلوبة'},
    {'key': 'workforce',    'title': 'القوى العاملة',    'icon': 'users-round',   'color': 'primary',  'description': 'توزيع الموظفين على الفروع والأقسام'},
    {'key': 'salary',       'title': 'الرواتب والمصاريف', 'icon': 'wallet',        'color': 'emerald',  'description': 'تحليل الرواتب والبدلات والاستقطاعات'},
    {'key': 'turnover',     'title': 'الدوران الوظيفي',   'icon': 'refresh-cw',    'color': 'indigo',   'description': 'التعيينات والإنهاءات ومعدل الدوران'},
    {'key': 'compliance',   'title': 'الالتزام والوثائق', 'icon': 'shield-check',  'color': 'rose',     'description': 'الوثائق والكروت الصحية والإنذارات'},
    {'key': 'leaves',       'title': 'الإجازات والغياب',  'icon': 'calendar-days', 'color': 'cyan',     'description': 'تقارير الإجازات والغياب'},
    {'key': 'demographics', 'title': 'تقارير ديموغرافية', 'icon': 'pie-chart',     'color': 'amber',    'description': 'توزيع حسب الجنس والجنسية والمهنة'},
    {'key': 'attendance',   'title': 'الحضور والبصمة',     'icon': 'fingerprint',   'color': 'violet',   'description': 'تقارير الحضور من أجهزة البصمة'},
]

_BASE_REPORTS = [
    {'group': 'workforce', 'key': 'headcount_summary',     'title': 'ملخص القوى العاملة',          'icon': 'users-round',   'color': 'primary',  'description': 'إجمالي الموظفين حسب الحالة والفرع'},
    {'group': 'workforce', 'key': 'branches',              'title': 'الموظفون حسب الفروع',         'icon': 'building-2',    'color': 'primary',  'description': 'توزيع الموظفين على الفروع'},
    {'group': 'workforce', 'key': 'departments_overview',  'title': 'الموظفون حسب الأقسام',        'icon': 'network',       'color': 'primary',  'description': 'توزيع الموظفين على الأقسام'},
    {'group': 'workforce', 'key': 'administrations_overview', 'title': 'الموظفون حسب الإدارات',   'icon': 'building',      'color': 'primary',  'description': 'توزيع الموظفين على الإدارات'},
    {'group': 'workforce', 'key': 'cost_centers_overview', 'title': 'الموظفون حسب مراكز التكلفة',  'icon': 'layers',        'color': 'primary',  'description': 'توزيع الموظفين والتكلفة'},
    {'group': 'salary',    'key': 'salary_expenses',       'title': 'تفاصيل الرواتب',              'icon': 'wallet',        'color': 'emerald',  'description': 'رواتب كل موظف بالتفصيل'},
    {'group': 'salary',    'key': 'allowances_breakdown',  'title': 'تفصيل البدلات',               'icon': 'plus-circle',   'color': 'emerald',  'description': 'بدلات كل موظف'},
    {'group': 'salary',    'key': 'deductions_breakdown',  'title': 'تفصيل الاستقطاعات',           'icon': 'minus-circle',  'color': 'emerald',  'description': 'استقطاعات آخر مسير'},
    {'group': 'salary',    'key': 'insurance_costs',       'title': 'بيانات التأمين',              'icon': 'shield',        'color': 'emerald',  'description': 'تأمين كل موظف'},
    {'group': 'turnover',  'key': 'new_hires',             'title': 'التعيينات الجديدة',           'icon': 'user-plus',     'color': 'indigo',   'description': 'الموظفون المعينون حديثاً'},
    {'group': 'turnover',  'key': 'terminations',          'title': 'انتهاء العقود',               'icon': 'user-minus',    'color': 'indigo',   'description': 'الموظفون المنتهية عقودهم والمصفون'},
    {'group': 'turnover',  'key': 'tenure_analysis',       'title': 'تحليل فترة الخدمة',           'icon': 'hourglass',     'color': 'indigo',   'description': 'مدة خدمة كل موظف'},
    {'group': 'compliance','key': 'health_cards',          'title': 'الكروت الصحية',               'icon': 'heart-pulse',   'color': 'rose',     'description': 'حالة الكروت الصحية'},
    {'group': 'compliance','key': 'warnings',              'title': 'الإنذارات والمخالفات',        'icon': 'alert-triangle','color': 'rose',     'description': 'الإنذارات والمخالفات'},
    {'group': 'leaves',    'key': 'leaves',                'title': 'سجل الإجازات',                'icon': 'plane',         'color': 'cyan',     'description': 'كل الإجازات المسجلة'},
    {'group': 'leaves',    'key': 'leave_balance',         'title': 'رصيد الإجازات',               'icon': 'calendar-clock','color': 'cyan',     'description': 'رصيد كل موظف'},
    {'group': 'leaves',    'key': 'absences',              'title': 'سجل الغياب',                  'icon': 'user-x',        'color': 'cyan',     'description': 'كل سجلات الغياب'},
    {'group': 'demographics','key': 'gender',              'title': 'حسب الجنس',                   'icon': 'users',         'color': 'amber',    'description': 'توزيع الموظفين حسب الجنس'},
    {'group': 'demographics','key': 'nationality',         'title': 'حسب الجنسية',                 'icon': 'flag',          'color': 'amber',    'description': 'توزيع حسب الجنسية'},
    {'group': 'demographics','key': 'professions',         'title': 'حسب المهنة',                  'icon': 'briefcase',     'color': 'amber',    'description': 'توزيع حسب المهنة'},
    {'group': 'attendance',  'key': 'biometric_daily',     'title': 'تقرير البصمة اليومي',         'icon': 'fingerprint',   'color': 'violet',   'description': 'دخول وخروج ومدة العمل لكل موظف/يوم'},
]

REPORTS = merge_reports_catalog(_BASE_REPORTS, PRIMARY_REPORT_SPECS)

MAX_REPORT_ROWS = 5000


def _materialize_qs(qs, max_rows: int = MAX_REPORT_ROWS):
    """تحميل صفوف من ORM بحد أقصى في SQL (استعلام +1 للكشف عن التجاوز)."""
    items = list(qs[: max_rows + 1])
    if len(items) > max_rows:
        return items[:max_rows], True
    return items, False


def _report_payload(columns, rows, truncated: bool = False, **extra) -> dict:
    data = {'columns': columns, 'rows': rows, **extra}
    if truncated:
        data['truncated'] = True
        data['max_rows'] = MAX_REPORT_ROWS
    return data


def _cap_report_data(data: dict) -> dict:
    """يحدّ عدد صفوف التقرير لحماية الذاكرة ووقت التصدير."""
    rows = data.get('rows') or []
    if len(rows) <= MAX_REPORT_ROWS:
        return data
    capped = dict(data)
    capped['rows'] = rows[:MAX_REPORT_ROWS]
    capped['truncated'] = True
    capped['total_rows'] = len(rows)
    capped['max_rows'] = MAX_REPORT_ROWS
    return capped

def _grouped_reports():
    groups = []
    for g in REPORT_GROUPS:
        items = [r for r in REPORTS if r['group'] == g['key']]
        if items:
            groups.append({**g, 'items': items})
    return groups


def _report_keys():
    return {r['key'] for r in REPORTS}


def _report_filters(request):
    """فلاتر موحّدة: فروع (متعدد)، كفالات (متعدد)، فترة زمنية."""
    today = date.today()
    first = today.replace(day=1)
    accessible = _user_accessible_branch_ids(request.user)
    branch_ids = parse_multi_filter_ids(request, 'branch', accessible_ids=accessible)
    if branch_ids is None and accessible is not None:
        branch_ids = list(accessible)
    sponsorship_ids = parse_multi_filter_ids(request, 'sponsorship')
    report = (request.GET.get('report') or '').strip()
    return {
        'branch_ids': branch_ids,
        'sponsorship_ids': sponsorship_ids,
        'date_from': request.GET.get('from') or first.isoformat(),
        'date_to': request.GET.get('to') or today.isoformat(),
        'report': report,
    }


def _parse_filter_dates(filters):
    df = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
    dt = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
    if df > dt:
        df, dt = dt, df
    return df, dt


def _report_filter_context(request):
    filters = _report_filters(request)
    branches_qs = Branch.objects.filter(is_deleted=False, is_active=True).order_by('name')
    branch_ids = _user_accessible_branch_ids(request.user)
    if branch_ids is not None:
        branches_qs = branches_qs.filter(pk__in=branch_ids)
    sponsorships = Sponsorship.objects.filter(is_deleted=False, is_active=True).order_by('company_name')
    return {
        'filter': filters,
        'branches': branches_qs,
        'sponsorships': sponsorships,
        'report_groups': _grouped_reports(),
    }


def _emp_qs():
    from apps.employees.models import Employee
    return Employee.objects.filter(is_deleted=False)

def _active():
    from apps.employees.models import Employee
    return _emp_qs().filter(status__in=[Employee.Status.ACTIVE, Employee.Status.LEAVE])


def _filtered_employees(request):
    """موظفون نشطون مع فلتر الفروع والكفالات."""
    filters = _report_filters(request)
    qs = _active()
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    return qs, filters


def _fmt_administration(employee) -> str:
    """عرض الإدارة: رقم — اسم أو — إن لم تُربط."""
    adm = getattr(employee, 'administration', None)
    if not adm:
        return '—'
    code = (getattr(adm, 'code', None) or '').strip()
    name = (getattr(adm, 'name', None) or '').strip()
    if code and name:
        return f'{code} — {name}'
    return code or name or '—'

# ══════════════════════════════════════════════════════════════════════════════
# دوال البناء — كل واحدة تُرجع columns + rows
# ══════════════════════════════════════════════════════════════════════════════

def _build_headcount_summary(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    cols = ['الاسم', 'الرقم الوظيفي', 'الفرع', 'القسم', 'الإدارة', 'الحالة', 'تاريخ المباشرة']
    labels = dict(Employee.Status.choices)
    qs = _emp_qs().select_related('branch', 'department', 'administration').order_by('branch__name', 'name')
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name, e.employee_number or '—',
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            labels.get(e.status, e.status),
            str(e.hire_date or '—'),
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)

def _build_branches(req):
    cols = ['الاسم', 'الرقم الوظيفي', 'الفرع', 'الأساسي', 'سكن', 'نقل', 'إضافي', 'كاش', 'تغذية', 'الإجمالي']
    qs = _filtered_employees(req)[0].select_related('branch').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.employee_number or '—', e.branch.name if e.branch else '—', str(e.basic_salary), str(e.housing_allowance), str(e.transport_allowance), str(e.other_allowance), str(e.cash_amount), str(e.meal_allowance), str(e.total_salary)] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_departments_overview(req):
    cols = ['الاسم', 'الفرع', 'القسم', 'الإدارة', 'مركز التكلفة', 'المسمى الوظيفي']
    qs = _filtered_employees(req)[0].select_related(
        'branch', 'department', 'administration', 'cost_center', 'profession',
    ).order_by('branch__name', 'department__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            e.cost_center.name if e.cost_center else '—',
            e.profession.name if e.profession else '—',
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)


def _build_administrations_overview(req):
    cols = ['الاسم', 'الفرع', 'رقم الإدارة', 'اسم الإدارة', 'القسم', 'مركز التكلفة']
    qs = _filtered_employees(req)[0].select_related(
        'branch', 'administration', 'department', 'cost_center',
    ).order_by('administration__code', 'administration__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = []
    for e in employees:
        adm = e.administration
        rows.append([
            e.name,
            e.branch.name if e.branch else '—',
            adm.code if adm else '—',
            adm.name if adm else '—',
            e.department.name if e.department else '—',
            e.cost_center.name if e.cost_center else '—',
        ])
    return _report_payload(cols, rows, truncated)


def _build_cost_centers_overview(req):
    cols = ['الاسم', 'مركز التكلفة', 'الفرع', 'القسم', 'الإدارة', 'الإجمالي']
    qs = _filtered_employees(req)[0].select_related(
        'branch', 'department', 'administration', 'cost_center',
    ).order_by('cost_center__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.cost_center.name if e.cost_center else '—',
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            str(e.total_salary),
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)

def _build_salary_expenses(req):
    cols = ['الاسم', 'الفرع', 'الأساسي', 'سكن', 'نقل', 'إضافي', 'كاش', 'تغذية', 'الإجمالي']
    qs = _filtered_employees(req)[0].select_related('branch').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', str(e.basic_salary), str(e.housing_allowance), str(e.transport_allowance), str(e.other_allowance), str(e.cash_amount), str(e.meal_allowance), str(e.total_salary)] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_allowances_breakdown(req):
    cols = ['الاسم', 'الفرع', 'سكن', 'نقل', 'إضافي', 'كاش', 'تغذية', 'إجمالي البدلات']
    qs = _filtered_employees(req)[0].select_related('branch').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = []
    for e in employees:
        t = e.housing_allowance + e.transport_allowance + e.other_allowance + e.cash_amount + e.meal_allowance
        rows.append([e.name, e.branch.name if e.branch else '—', str(e.housing_allowance), str(e.transport_allowance), str(e.other_allowance), str(e.cash_amount), str(e.meal_allowance), str(t)])
    return _report_payload(cols, rows, truncated)

def _build_deductions_breakdown(req):
    from apps.payroll.models import PayrollRun
    last = PayrollRun.objects.filter(status=PayrollRun.Status.LOCKED).order_by('-period_year', '-period_month').first()
    cols = ['الموظف', 'غياب', 'إجازة بدون راتب', 'سلف', 'مخالفات', 'تأمينات', 'أخرى', 'إجمالي الخصم']
    if not last:
        return {'columns': cols, 'rows': [], 'note': 'لا يوجد مسير مُرحَّل'}
    lines_qs = last.lines.select_related('employee').order_by('employee__name')
    line_items, truncated = _materialize_qs(lines_qs)
    rows = [[l.employee.name, str(l.absence_deduction), str(l.unpaid_leave_deduction), str(l.loan_deduction), str(l.penalty_deduction), str(l.insurance_deduction), str(l.other_deduction), str(l.total_deductions)] for l in line_items]
    return _report_payload(cols, rows, truncated, note=f'من مسير: {last}')

def _build_insurance_costs(req):
    cols = ['الاسم', 'الفرع', 'شركة التأمين', 'فئة التأمين', 'نسبة الخصم %']
    qs = _filtered_employees(req)[0].select_related('branch', 'insurance', 'insurance_class').order_by('insurance__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', e.insurance.name if e.insurance else '—', e.insurance_class.name if e.insurance_class else '—', str(e.insurance_deduction_rate)] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_new_hires(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الاسم', 'الرقم الوظيفي', 'الفرع', 'القسم', 'الإدارة', 'تاريخ المباشرة', 'الجنسية']
    qs = _emp_qs().filter(hire_date__gte=df, hire_date__lte=dt).exclude(status=Employee.Status.TERMINATED)
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.select_related('branch', 'department', 'administration', 'nationality').order_by('-hire_date')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name, e.employee_number or '—',
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            str(e.hire_date or '—'),
            e.nationality.name if e.nationality else '—',
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated, note=f'تعيينات من {df} إلى {dt}')

def _build_terminations(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الاسم', 'الفرع', 'تاريخ المباشرة', 'تاريخ الانتهاء', 'السبب', 'إجمالي الراتب الأخير']
    qs = _emp_qs().filter(status=Employee.Status.TERMINATED, end_date__gte=df, end_date__lte=dt)
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.select_related('branch').order_by('-end_date')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', str(e.hire_date or '—'), str(e.end_date or '—'), e.end_reason or '—', str(e.total_salary)] for e in employees]
    return _report_payload(cols, rows, truncated, note=f'تصفيات من {df} إلى {dt}')

def _build_tenure_analysis(req):
    today = date.today()
    cols = ['الاسم', 'الفرع', 'تاريخ المباشرة', 'مدة الخدمة (سنة)', 'مدة الخدمة (يوم)']
    qs = _filtered_employees(req)[0].exclude(hire_date__isnull=True).select_related('branch').order_by('hire_date')
    employees, truncated = _materialize_qs(qs)
    rows = []
    for e in employees:
        days = (today - e.hire_date).days
        years = round(days / 365.25, 1)
        rows.append([e.name, e.branch.name if e.branch else '—', str(e.hire_date), str(years), str(days)])
    return _report_payload(cols, rows, truncated)

def _build_health_cards(req):
    today = date.today()
    soon = today + timedelta(days=90)
    cols = ['الاسم', 'الفرع', 'حالة الكرت', 'تاريخ الانتهاء', 'الوضع']
    qs = _filtered_employees(req)[0].select_related('branch').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    labels = {'available': 'متوفر', 'not_available': 'غير متوفر'}
    rows = []
    for e in employees:
        st = labels.get(e.health_card_status, e.health_card_status)
        exp = str(e.health_card_expiry) if e.health_card_expiry else '—'
        if not e.health_card_expiry:
            flag = '—'
        elif e.health_card_expiry < today:
            flag = '❌ منتهي'
        elif e.health_card_expiry <= soon:
            flag = '⚠️ ينتهي قريباً'
        else:
            flag = '✅ ساري'
        rows.append([e.name, e.branch.name if e.branch else '—', st, exp, flag])
    return _report_payload(cols, rows, truncated)

def _build_warnings(req):
    from apps.employees.models import EmployeeStatement
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الموظف', 'الفرع', 'النوع', 'العنوان', 'التاريخ', 'مبلغ الخصم']
    types = [EmployeeStatement.StatementType.WARNING, EmployeeStatement.StatementType.FINAL_WARNING, EmployeeStatement.StatementType.PENALTY]
    qs = EmployeeStatement.objects.filter(
        statement_type__in=types,
        is_deleted=False,
        statement_date__gte=df,
        statement_date__lte=dt,
    ).select_related('employee', 'employee__branch')
    qs = apply_branch_filter(qs, filters['branch_ids'], field='employee__branch_id')
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.order_by('-statement_date')[:500]
    labels = dict(EmployeeStatement.StatementType.choices)
    rows = [[s.employee.name, s.employee.branch.name if s.employee.branch else '—', labels.get(s.statement_type, s.statement_type), s.title, str(s.statement_date), str(s.deduction_amount)] for s in qs]
    return {'columns': cols, 'rows': rows}

def _build_leaves(req):
    from apps.employees.models import EmployeeLeave
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الموظف', 'الفرع', 'نوع الإجازة', 'من', 'إلى', 'عدد الأيام']
    labels = dict(EmployeeLeave.LeaveType.choices)
    qs = EmployeeLeave.objects.filter(
        is_deleted=False,
        date_from__lte=dt,
        date_to__gte=df,
    ).select_related('employee', 'employee__branch')
    qs = apply_branch_filter(qs, filters['branch_ids'], field='employee__branch_id')
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.order_by('-date_from')[:500]
    rows = [[l.employee.name, l.employee.branch.name if l.employee.branch else '—', labels.get(l.leave_type, l.leave_type), str(l.date_from), str(l.date_to), str(l.days)] for l in qs]
    return {'columns': cols, 'rows': rows}

def _build_leave_balance(req):
    cols = ['الاسم', 'الفرع', 'تاريخ المباشرة', 'المستحق', 'المستخدم', 'المتبقي']
    qs = _filtered_employees(req)[0].exclude(hire_date__isnull=True).exclude(sponsorship__isnull=True).select_related('branch').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', str(e.hire_date), str(e.accrued_leave_days), str(e.available_leave_balance), str(e.remaining_leave_days)] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_biometric_daily(req):
    from apps.attendance.selectors.biometric_devices import filter_biometric_devices_for_user
    from apps.attendance.selectors.daily_report import build_daily_attendance_rows, daily_rows_to_table
    from apps.attendance.selectors.punch_records import get_punch_queryset
    from apps.core.utils.attendance_filters import clamp_attendance_date_range

    filters = _report_filters(req)
    report_filters = {
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
    }
    report_filters, date_clamped = clamp_attendance_date_range(report_filters)
    date_from, date_to = _parse_filter_dates({**filters, **report_filters})
    qs = get_punch_queryset(
        branch_ids=filters['branch_ids'],
        date_from=date_from,
        date_to=date_to,
    ).filter(device_id__in=filter_biometric_devices_for_user(req.user).values('pk'))
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    daily_rows = build_daily_attendance_rows(qs)
    if len(daily_rows) > MAX_REPORT_ROWS:
        daily_rows = daily_rows[:MAX_REPORT_ROWS]
        truncated = True
    else:
        truncated = False
    data = daily_rows_to_table(daily_rows)
    note = f'من {date_from} إلى {date_to}'
    if date_clamped:
        note += ' — تم تقييد الفترة إلى 93 يوماً'
    if truncated:
        note += f' — عُرض أول {MAX_REPORT_ROWS} صف'
    data['note'] = note + ' — للفلترة الكاملة: قائمة البصمة → تقرير البصمة'
    if truncated:
        data['truncated'] = True
        data['max_rows'] = MAX_REPORT_ROWS
    return data


def _build_absences(req):
    from apps.employees.models import EmployeeAbsence
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الموظف', 'الفرع', 'تاريخ الغياب', 'عدد الأيام', 'سبب الغياب', 'مبلغ الخصم', 'محتسب في مسير']
    qs = EmployeeAbsence.objects.filter(
        is_deleted=False,
        absence_date__gte=df,
        absence_date__lte=dt,
    ).select_related('employee', 'employee__branch', 'applied_to_payroll')
    qs = apply_branch_filter(qs, filters['branch_ids'], field='employee__branch_id')
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.order_by('-absence_date')[:500]
    rows = [
        [
            a.employee.name,
            a.employee.branch.name if a.employee.branch else '—',
            str(a.absence_date),
            str(a.days),
            a.reason or a.notes or '—',
            str(a.deduction_amount),
            str(a.applied_to_payroll or '—'),
        ]
        for a in qs
    ]
    return {'columns': cols, 'rows': rows}


def _build_employees(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    cols = ['الاسم', 'الرقم الوظيفي', 'الفرع', 'القسم', 'الإدارة', 'الحالة', 'تاريخ المباشرة', 'الجوال']
    labels = dict(Employee.Status.choices)
    qs = _emp_qs().select_related('branch', 'department', 'administration').order_by('branch__name', 'name')
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.employee_number or '—',
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            labels.get(e.status, e.status),
            str(e.hire_date or '—'),
            e.phone or '—',
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)


def _build_stopped(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الاسم', 'الفرع', 'تاريخ المباشرة', 'تاريخ التوقف', 'سبب التوقف', 'الحالة']
    labels = dict(Employee.Status.choices)
    qs = _emp_qs().filter(status=Employee.Status.TERMINATED)
    qs = qs.filter(
        Q(end_date__gte=df, end_date__lte=dt) | Q(end_date__isnull=True),
    )
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.select_related('branch').order_by('-end_date', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.branch.name if e.branch else '—',
            str(e.hire_date or '—'),
            str(e.end_date or '—'),
            e.end_reason or '—',
            labels.get(e.status, e.status),
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated, note=f'متوقفون — من {df} إلى {dt}')


def _build_statements(req):
    from apps.employees.models import EmployeeStatement
    filters = _report_filters(req)
    df, dt = _parse_filter_dates(filters)
    cols = ['الموظف', 'الفرع', 'النوع', 'العنوان', 'التاريخ', 'الرقم المتسلسل']
    types = [
        EmployeeStatement.StatementType.STATEMENT,
        EmployeeStatement.StatementType.ACKNOWLEDGMENT,
        EmployeeStatement.StatementType.OTHER,
    ]
    qs = EmployeeStatement.objects.filter(
        statement_type__in=types,
        is_deleted=False,
        statement_date__gte=df,
        statement_date__lte=dt,
    ).select_related('employee', 'employee__branch')
    qs = apply_branch_filter(qs, filters['branch_ids'], field='employee__branch_id')
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.order_by('-statement_date')[:500]
    labels = dict(EmployeeStatement.StatementType.choices)
    rows = [
        [
            s.employee.name,
            s.employee.branch.name if s.employee.branch else '—',
            labels.get(s.statement_type, s.statement_type),
            s.title,
            str(s.statement_date),
            s.serial_number or '—',
        ]
        for s in qs
    ]
    return {'columns': cols, 'rows': rows}


def _build_housing(req):
    cols = ['الاسم', 'الفرع', 'السكن', 'القسم', 'الجوال']
    qs = _filtered_employees(req)[0].select_related('branch', 'housing', 'department').order_by(
        'housing__name', 'branch__name', 'name',
    )
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.branch.name if e.branch else '—',
            e.housing.name if e.housing else '—',
            e.department.name if e.department else '—',
            e.phone or '—',
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)


def _build_active_headcount(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    cols = ['الاسم', 'الرقم الوظيفي', 'الفرع', 'القسم', 'الإدارة', 'تاريخ المباشرة']
    qs = _emp_qs().filter(status=Employee.Status.ACTIVE)
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.select_related('branch', 'department', 'administration').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.employee_number or '—',
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            _fmt_administration(e),
            str(e.hire_date or '—'),
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)


def _build_suspended(req):
    from apps.employees.models import Employee
    filters = _report_filters(req)
    cols = ['الاسم', 'الفرع', 'القسم', 'تاريخ المباشرة', 'الحالة', 'الجوال']
    labels = dict(Employee.Status.choices)
    qs = _emp_qs().filter(status=Employee.Status.SUSPENDED)
    qs = apply_branch_filter(qs, filters['branch_ids'])
    if filters['sponsorship_ids']:
        qs = qs.filter(sponsorship_id__in=filters['sponsorship_ids'])
    qs = qs.select_related('branch', 'department').order_by('branch__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [
        [
            e.name,
            e.branch.name if e.branch else '—',
            e.department.name if e.department else '—',
            str(e.hire_date or '—'),
            labels.get(e.status, e.status),
            e.phone or '—',
        ]
        for e in employees
    ]
    return _report_payload(cols, rows, truncated)


def _build_attendance_late(req):
    from django.utils import timezone
    from apps.attendance.models import EmployeeBiometricSettings
    from apps.attendance.selectors.biometric_devices import filter_biometric_devices_for_user
    from apps.attendance.selectors.daily_report import build_daily_attendance_rows
    from apps.attendance.selectors.punch_records import get_punch_queryset
    from apps.attendance.services.attendance_evaluation import (
        evaluate_daily_checkin,
        evaluate_daily_checkout,
    )
    from apps.core.utils.attendance_filters import clamp_attendance_date_range

    filters = _report_filters(req)
    report_filters, date_clamped = clamp_attendance_date_range({
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
    })
    date_from, date_to = _parse_filter_dates({**filters, **report_filters})
    qs = get_punch_queryset(
        branch_ids=filters['branch_ids'],
        date_from=date_from,
        date_to=date_to,
    ).filter(device_id__in=filter_biometric_devices_for_user(req.user).values('pk'))
    if filters['sponsorship_ids']:
        qs = qs.filter(employee__sponsorship_id__in=filters['sponsorship_ids'])
    daily_rows = build_daily_attendance_rows(qs)
    rows_truncated = len(daily_rows) > MAX_REPORT_ROWS
    if rows_truncated:
        daily_rows = daily_rows[:MAX_REPORT_ROWS]
    employee_ids = [r.employee_id for r in daily_rows if r.employee_id]
    settings_map = {
        s.employee_id: s
        for s in EmployeeBiometricSettings.objects.filter(employee_id__in=employee_ids)
    }

    cols = [
        'التاريخ', 'الموظف', 'الفرع', 'الدخول المتوقع', 'وقت الدخول',
        'تأخير الدخول (د)', 'الخروج المتوقع', 'وقت الخروج', 'خروج مبكر (د)', 'الملاحظة',
    ]
    table_rows = []

    for row in daily_rows:
        if not row.employee_id or not row.is_mapped:
            continue
        settings = settings_map.get(row.employee_id)
        if not settings:
            continue

        late_in = ''
        early_out = ''
        notes = []

        exp_in = settings.expected_check_in.strftime('%H:%M') if settings.expected_check_in else '—'
        exp_out = settings.expected_check_out.strftime('%H:%M') if settings.expected_check_out else '—'
        act_in = timezone.localtime(row.check_in).strftime('%H:%M') if row.check_in else '—'
        act_out = timezone.localtime(row.check_out).strftime('%H:%M') if row.check_out else '—'

        checkin_eval = evaluate_daily_checkin(row.work_date, row.check_in, settings)
        if checkin_eval and checkin_eval.is_late:
            late_in = str(checkin_eval.late_minutes)
            notes.append(f'تأخر دخول {checkin_eval.late_minutes} د')

        checkout_eval = evaluate_daily_checkout(row.work_date, row.check_out, settings)
        if checkout_eval and checkout_eval.is_early:
            early_out = str(checkout_eval.early_minutes)
            notes.append(f'خروج مبكر {checkout_eval.early_minutes} د')

        if not notes:
            continue

        table_rows.append([
            str(row.work_date),
            row.employee_name,
            row.branch_name,
            exp_in,
            act_in,
            late_in or '—',
            exp_out,
            act_out,
            early_out or '—',
            ' · '.join(notes),
        ])

    note = f'من {date_from} إلى {date_to} — يظهر فقط من لديه إعدادات بصمة وتأخر فعلي'
    if date_clamped:
        note += ' — تم تقييد الفترة إلى 93 يوماً'
    if rows_truncated:
        note += f' — عُرض أول {MAX_REPORT_ROWS} يوم-موظف'
    data = {'columns': cols, 'rows': table_rows, 'note': note}
    if rows_truncated:
        data['truncated'] = True
        data['max_rows'] = MAX_REPORT_ROWS
    return data

def _build_gender(req):
    from apps.employees.models import Employee
    cols = ['الاسم', 'الفرع', 'الجنس', 'الجنسية', 'المهنة']
    labels = dict(Employee.Gender.choices)
    qs = _filtered_employees(req)[0].select_related('branch', 'nationality', 'profession').order_by('gender', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', labels.get(e.gender, e.gender or 'غير محدد'), e.nationality.name if e.nationality else '—', e.profession.name if e.profession else '—'] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_nationality(req):
    cols = ['الاسم', 'الفرع', 'الجنسية', 'رقم الهوية', 'رقم الجوال']
    qs = _filtered_employees(req)[0].select_related('branch', 'nationality').order_by('nationality__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', e.nationality.name if e.nationality else '—', e.id_number or '—', e.phone or '—'] for e in employees]
    return _report_payload(cols, rows, truncated)

def _build_professions(req):
    cols = ['الاسم', 'الفرع', 'المهنة', 'الجنسية', 'الراتب الإجمالي']
    qs = _filtered_employees(req)[0].select_related('branch', 'profession', 'nationality').order_by('profession__name', 'name')
    employees, truncated = _materialize_qs(qs)
    rows = [[e.name, e.branch.name if e.branch else '—', e.profession.name if e.profession else '—', e.nationality.name if e.nationality else '—', str(e.total_salary)] for e in employees]
    return _report_payload(cols, rows, truncated)

BUILDERS = {
    'headcount_summary': _build_headcount_summary, 'branches': _build_branches,
    'departments_overview': _build_departments_overview,
    'administrations_overview': _build_administrations_overview,
    'cost_centers_overview': _build_cost_centers_overview,
    'salary_expenses': _build_salary_expenses, 'allowances_breakdown': _build_allowances_breakdown,
    'deductions_breakdown': _build_deductions_breakdown, 'insurance_costs': _build_insurance_costs,
    'new_hires': _build_new_hires, 'terminations': _build_terminations, 'tenure_analysis': _build_tenure_analysis,
    'health_cards': _build_health_cards, 'warnings': _build_warnings,
    'leaves': _build_leaves, 'leave_balance': _build_leave_balance, 'absences': _build_absences,
    'gender': _build_gender, 'nationality': _build_nationality, 'professions': _build_professions,
    'biometric_daily': _build_biometric_daily,
    'employees': _build_employees,
    'stopped': _build_stopped,
    'statements': _build_statements,
    'housing': _build_housing,
    'active_headcount': _build_active_headcount,
    'suspended': _build_suspended,
    'attendance_late': _build_attendance_late,
}

def _catalog_for_user(user):
    """تقارير ومجموعات مرئية حسب صلاحيات الرواتب."""
    reports = [r for r in REPORTS if report_allowed_for_user(user, r['key'])]
    groups = [
        g for g in REPORT_GROUPS
        if g['key'] != 'salary' or user_can_view_financial_reports(user)
    ]
    return reports, groups


def _grouped_reports_for_user(user):
    """مجموعات التقارير مع items[] لكل مجموعة — للقوائم والفلاتر."""
    visible_reports, visible_groups = _catalog_for_user(user)
    grouped = []
    for g in visible_groups:
        items = [r for r in visible_reports if r.get('group') == g['key']]
        if items:
            grouped.append({**g, 'items': items})
    return grouped


def _filter_querystring(request, exclude=()):
    f = _report_filters(request)
    params: list[tuple[str, object]] = []
    if 'branch' not in exclude:
        append_multi_param(params, 'branch', f.get('branch_ids'))
    if 'sponsorship' not in exclude:
        append_multi_param(params, 'sponsorship', f.get('sponsorship_ids'))
    for key, val in (
        ('from', f['date_from']),
        ('to', f['date_to']),
    ):
        if key in exclude or not val:
            continue
        params.append((key, val))
    report = f.get('report')
    if 'report' not in exclude and report:
        params.append(('report', report))
    return urlencode(params, doseq=True)


@login_required
@permission_required('reports.view')
def reports_index(request):
    visible_reports, visible_groups = _catalog_for_user(request.user)
    visible_keys = {r['key'] for r in visible_reports}
    new_report = (request.GET.get('report') or '').strip()
    if new_report and new_report in visible_keys:
        qs = _filter_querystring(request, exclude=('report',))
        url = reverse('web:report_detail', kwargs={'report_type': new_report})
        if qs:
            url = f'{url}?{qs}'
        return redirect(url)

    ctx = _report_filter_context(request)
    return render(request, 'pages/reports/index.html', {
        **ctx,
        'reports': visible_reports,
        'report_groups': _grouped_reports_for_user(request.user),
        'clear_url': reverse('web:reports_index'),
    })

@login_required
@permission_required('reports.view')
def multi_report_detail(request):
    visible_reports, _ = _catalog_for_user(request.user)
    visible_by_key = {r['key']: r for r in visible_reports}
    report_keys = request.GET.get('reports', '').split(',')
    selected_reports = []
    
    for key in report_keys:
        key = key.strip()
        if not key:
            continue
        meta = visible_by_key.get(key)
        if meta:
            builder = BUILDERS.get(key)
            data = builder(request) if builder else {'columns': [], 'rows': []}
            selected_reports.append({
                'meta': meta,
                'data': _cap_report_data(data),
            })
            
    if not selected_reports:
        messages.info(request, 'اختر تقاريراً لعرضها من صفحة التقارير.')
        return redirect('web:reports_index')

    return render(request, 'pages/reports/multi_detail.html', {
        'reports_data': selected_reports,
        'reports_count': len(selected_reports)
    })

@login_required
@permission_required('reports.view')
def report_detail(request, report_type):
    visible_reports, visible_groups = _catalog_for_user(request.user)
    visible_keys = {r['key'] for r in visible_reports}
    meta = next((r for r in visible_reports if r['key'] == report_type), None)
    if not meta:
        if report_type in {r['key'] for r in REPORTS}:
            messages.error(request, 'لا تملك صلاحية عرض هذا التقرير (بيانات رواتب).')
            return redirect('web:reports_index')
        raise Http404("تقرير غير معروف")

    new_report = (request.GET.get('report') or '').strip()
    if new_report and new_report != report_type and new_report in visible_keys:
        qs = _filter_querystring(request, exclude=('report',))
        url = reverse('web:report_detail', kwargs={'report_type': new_report})
        if qs:
            url = f'{url}?{qs}'
        return redirect(url)

    group = next((g for g in visible_groups if g['key'] == meta.get('group')), None)
    builder = BUILDERS.get(report_type)
    filters = _report_filters(request)
    from apps.core.services.report_cache import cache_bypass_requested, get_or_build_report_data

    def _build():
        raw = builder(request) if builder else {'columns': [], 'rows': []}
        return _cap_report_data(raw)

    data, from_cache = get_or_build_report_data(
        user_id=request.user.id,
        report_type=report_type,
        filters=filters,
        builder=_build,
        bypass=cache_bypass_requested(request),
    )
    ctx = _report_filter_context(request)
    return render(request, 'pages/reports/detail.html', {
        'report_meta': meta,
        'group_meta': group,
        'reports': visible_reports,
        'report_groups': _grouped_reports_for_user(request.user),
        'data': data,
        'report_type': report_type,
        'selected_report': report_type,
        'form_action': reverse('web:report_detail', kwargs={'report_type': report_type}),
        'clear_url': reverse('web:report_detail', kwargs={'report_type': report_type}),
        'export_url': reverse('web:report_export_excel', kwargs={'report_type': report_type}),
        'filter_querystring': _filter_querystring(request),
        'from_cache': from_cache,
        **ctx,
    })


@login_required
@any_permission_required('reports.export')
def report_export_excel(request, report_type):
    """تصدير تقرير إلى Excel — يستخدم نفس بيانات العرض (مع كاش)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة.')
        return redirect('web:report_detail', report_type=report_type)

    visible_reports, _ = _catalog_for_user(request.user)
    meta = next((r for r in visible_reports if r['key'] == report_type), None)
    if not meta:
        if report_type in {r['key'] for r in REPORTS}:
            messages.error(request, 'لا تملك صلاحية تصدير هذا التقرير (بيانات رواتب).')
            return redirect('web:reports_index')
        raise Http404('تقرير غير معروف')

    builder = BUILDERS.get(report_type)
    filters = _report_filters(request)
    from apps.core.services.report_cache import cache_bypass_requested, get_or_build_report_data

    def _build():
        return _cap_report_data(builder(request) if builder else {'columns': [], 'rows': []})

    data, _ = get_or_build_report_data(
        user_id=request.user.id,
        report_type=report_type,
        filters=filters,
        builder=_build,
        bypass=cache_bypass_requested(request),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = (meta.get('title') or report_type)[:31]
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill('solid', fgColor='1E40AF')
    columns = data.get('columns') or []
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row in enumerate(data.get('rows') or [], 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    stamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'report_{report_type}_{stamp}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
