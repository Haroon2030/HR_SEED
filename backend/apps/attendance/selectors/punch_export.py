"""تحويل سجلات البصمة لجدول العرض والتصدير — أعمدة موحّدة."""
from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from django.http import HttpResponse
from django.utils import timezone

from apps.attendance.models import AttendancePunch

EXPORT_MAX_ROWS = 50_000

if TYPE_CHECKING:
    from apps.attendance.services.attendance_pull import EnrichedPunch

PUNCH_TABLE_COLUMNS = [
    '#',
    'التاريخ',
    'الوقت',
    'معرف ZK',
    'رقم المستخدم',
    'الاسم على الجهاز',
    'موظف HR',
    'الرقم الوظيفي',
    'نوع الحركة',
    'St',
    'التحقق',
    'Vm',
    'الجهاز',
    'IP',
]


def format_punch_table_row(
    *,
    index: int,
    punched_at,
    device_record_uid,
    device_user_id,
    device_user_name: str,
    employee_name: str,
    employee_number: str,
    punch_type_label: str,
    raw_status,
    verify_mode_label: str,
    verify_mode,
    device_name: str,
    device_ip: str,
) -> list:
    local = timezone.localtime(punched_at)
    return [
        index,
        local.strftime('%Y-%m-%d'),
        local.strftime('%H:%M:%S'),
        str(device_record_uid or '—'),
        str(device_user_id),
        device_user_name or '—',
        employee_name or 'غير مربوط',
        employee_number or '—',
        punch_type_label or '—',
        str(raw_status if raw_status is not None else '—'),
        verify_mode_label or '—',
        str(verify_mode if verify_mode is not None else '—'),
        device_name or '—',
        device_ip or '—',
    ]


def attendance_punch_to_row(p: AttendancePunch, index: int) -> list:
    return format_punch_table_row(
        index=index,
        punched_at=p.punched_at,
        device_record_uid=p.device_record_uid,
        device_user_id=p.device_user_id,
        device_user_name=p.device_user_name or '',
        employee_name=p.employee.name if p.employee else '',
        employee_number=p.employee.employee_number if p.employee else '',
        punch_type_label=p.get_punch_type_display(),
        raw_status=p.raw_status,
        verify_mode_label=p.verify_mode_label or '',
        verify_mode=p.verify_mode,
        device_name=p.device.name,
        device_ip=p.device.ip_address,
    )


def enriched_punch_to_row(
    p: EnrichedPunch,
    index: int,
    *,
    device_name: str,
    device_ip: str,
) -> list:
    return format_punch_table_row(
        index=index,
        punched_at=p.punched_at,
        device_record_uid=p.device_record_uid,
        device_user_id=p.device_user_id,
        device_user_name=p.device_user_name or '',
        employee_name=p.employee_name or '',
        employee_number=p.employee_number or '',
        punch_type_label=p.punch_type_label,
        raw_status=p.raw_status,
        verify_mode_label=p.verify_mode_label or '',
        verify_mode=p.verify_mode,
        device_name=device_name,
        device_ip=device_ip,
    )


def punches_to_table_rows(qs, *, start_index: int = 1, max_rows: int | None = EXPORT_MAX_ROWS) -> dict:
    rows = []
    for offset, p in enumerate(qs.iterator(chunk_size=5000)):
        if max_rows is not None and offset >= max_rows:
            break
        rows.append(attendance_punch_to_row(p, start_index + offset))
    return {'columns': PUNCH_TABLE_COLUMNS, 'rows': rows}


def enriched_punches_to_table_rows(
    punches: list,
    *,
    device_name: str,
    device_ip: str,
    start_index: int = 1,
) -> dict:
    rows = [
        enriched_punch_to_row(
            p,
            start_index + offset,
            device_name=device_name,
            device_ip=device_ip,
        )
        for offset, p in enumerate(punches)
    ]
    return {'columns': PUNCH_TABLE_COLUMNS, 'rows': rows}


def write_punch_table_sheet(ws, table: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill('solid', fgColor='1E40AF')
    for col, header in enumerate(table['columns'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row in enumerate(table['rows'], 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    for col in range(1, len(table['columns']) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def punch_table_http_response(
    table: dict,
    *,
    filename_prefix: str,
    sheet_title: str = 'سجلات الحضور',
) -> HttpResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    write_punch_table_sheet(ws, table)
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{stamp}.xlsx"'
    return response
