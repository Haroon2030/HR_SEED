"""
سحب احترافي لسجلات الحضور من أجهزة ZKTeco.

يُستخدم من:
  - python manage.py pull_biometric_attendance
  - الويب (أجهزة عامة) أو وكيل الفرع (192.168.x.x)
"""
from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import TYPE_CHECKING

from django.utils import timezone

from apps.attendance.services.labels import punch_type_for_status, verify_mode_label as zk_verify_label
from apps.attendance.services.zk_client import (
    DeviceSnapshot,
    DeviceUserRow,
    RawAttendanceRow,
    fetch_device_snapshot,
    sync_device_users,
)

if TYPE_CHECKING:
    from apps.attendance.models import BiometricDevice


@dataclass
class EnrichedPunch:
    device_user_id: int
    device_user_name: str
    punched_at: datetime
    punch_type: str
    punch_type_label: str
    verify_mode: int | None
    verify_mode_label: str
    device_record_uid: int | None
    raw_status: int | None
    employee_id: int | None = None
    employee_name: str = ''
    employee_number: str = ''


@dataclass
class DevicePullResult:
    device_id: int
    device_name: str
    device_address: str
    ok: bool
    device_ip: str = ''
    error: str = ''
    serial_number: str = ''
    firmware: str = ''
    users_on_device: int = 0
    punches_fetched: int = 0
    punches_after_filter: int = 0
    punches_new: int = 0
    skipped_time_filter: int = 0
    imported: int = 0
    skipped_duplicate: int = 0
    incremental: bool = True
    unmapped_users: int = 0
    date_from: date | None = None
    date_to: date | None = None
    export_path: str = ''
    batch: str = ''
    punches: list[EnrichedPunch] = field(default_factory=list)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(value)


def _filter_rows(
    rows: list[RawAttendanceRow],
    date_from: date | None,
    date_to: date | None,
) -> list[RawAttendanceRow]:
    if not date_from and not date_to:
        return rows
    filtered: list[RawAttendanceRow] = []
    for row in rows:
        local_date = timezone.localtime(row.punched_at).date()
        if date_from and local_date < date_from:
            continue
        if date_to and local_date > date_to:
            continue
        filtered.append(row)
    return filtered


def _build_name_map(users: list[DeviceUserRow]) -> dict[int, str]:
    return {u.device_user_id: u.name for u in users if u.name}


def _enrich_punches(
    device: BiometricDevice,
    rows: list[RawAttendanceRow],
    name_map: dict[int, str],
    enroll_map: dict[int, tuple[int, str, str]],
) -> list[EnrichedPunch]:
    from apps.attendance.models import AttendancePunch

    type_labels = dict(AttendancePunch.PunchType.choices)
    enriched: list[EnrichedPunch] = []
    for row in sorted(rows, key=lambda r: r.punched_at):
        punch_code = row.punch_type
        _, punch_label = punch_type_for_status(row.status)
        punch_label = type_labels.get(punch_code, punch_label)

        emp_id, emp_name, emp_num = enroll_map.get(row.device_user_id, (None, '', ''))
        enriched.append(
            EnrichedPunch(
                device_user_id=row.device_user_id,
                device_user_name=name_map.get(row.device_user_id, ''),
                punched_at=row.punched_at,
                punch_type=punch_code,
                punch_type_label=punch_label,
                verify_mode=row.verify_mode,
                verify_mode_label=zk_verify_label(row.verify_mode),
                device_record_uid=row.uid,
                raw_status=row.status,
                employee_id=emp_id,
                employee_name=emp_name,
                employee_number=emp_num,
            )
        )
    return enriched


def _get_enrollment_map(device: BiometricDevice) -> dict[int, tuple[int, str, str]]:
    from apps.attendance.models import EmployeeBiometricEnrollment

    result: dict[int, tuple[int, str, str]] = {}
    qs = (
        EmployeeBiometricEnrollment.objects.filter(device=device, is_deleted=False)
        .select_related('employee')
    )
    for en in qs:
        emp = en.employee
        result[en.device_user_id] = (
            emp.id,
            emp.name,
            emp.employee_number or '',
        )
    return result


def _persist_users(device: BiometricDevice, users: list[DeviceUserRow]) -> None:
    from apps.attendance.models import BiometricDeviceUser

    now = timezone.now()
    for u in users:
        BiometricDeviceUser.objects.update_or_create(
            device=device,
            device_user_id=u.device_user_id,
            defaults={
                'name': u.name,
                'card': u.card,
                'privilege': u.privilege,
                'last_synced_at': now,
            },
        )


def export_punches_excel(
    result: DevicePullResult,
    output_path: Path,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'ملخص'
    ws_summary.sheet_view.rightToLeft = True

    header_fill = PatternFill('solid', fgColor='1E40AF')
    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    summary_rows = [
        ('الجهاز', result.device_name),
        ('العنوان', result.device_address),
        ('الرقم التسلسلي', result.serial_number or '—'),
        ('الإصدار', result.firmware or '—'),
        ('المستخدمون على الجهاز', result.users_on_device),
        ('سجلات على الجهاز', result.punches_fetched),
        ('بعد التصفية', result.punches_after_filter),
        ('جديد للاستيراد', result.punches_new),
        ('مستوردة للنظام', result.imported),
        ('مكررة (تخطي)', result.skipped_duplicate),
        ('قديمة (تخطي زمني)', result.skipped_time_filter),
        ('غير مربوطين بـ HR', result.unmapped_users),
        ('من تاريخ', str(result.date_from or 'الكل')),
        ('إلى تاريخ', str(result.date_to or 'الكل')),
        ('ملف التصدير', str(output_path.name)),
        ('دفعة المزامنة', result.batch or '—'),
    ]
    ws_summary['A1'] = 'تقرير سحب الحضور — ZKTeco'
    ws_summary['A1'].font = Font(bold=True, size=14, name='Arial')
    ws_summary.merge_cells('A1:B1')
    for i, (label, val) in enumerate(summary_rows, start=3):
        ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True, name='Arial')
        ws_summary.cell(row=i, column=2, value=val)

    from apps.attendance.selectors.punch_export import (
        enriched_punches_to_table_rows,
        write_punch_table_sheet,
    )

    ws = wb.create_sheet('سجلات الحضور')
    punch_table = enriched_punches_to_table_rows(
        result.punches,
        device_name=result.device_name,
        device_ip=result.device_ip,
    )
    write_punch_table_sheet(ws, punch_table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def pull_device_attendance(
    device: BiometricDevice,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    import_db: bool = True,
    dry_run: bool = False,
    clear_device: bool = False,
    incremental: bool | None = None,
    force_mock: bool | None = None,
    export_path: Path | None = None,
) -> DevicePullResult:
    use_incremental = incremental if incremental is not None else not (date_from or date_to)

    result = DevicePullResult(
        device_id=device.id,
        device_name=device.name,
        device_address=device.address_label,
        device_ip=device.ip_address or '',
        ok=False,
        date_from=date_from,
        date_to=date_to,
        incremental=use_incremental,
    )

    snapshot, error = fetch_device_snapshot(
        device, clear_after=clear_device, force_mock=force_mock,
    )
    if error or snapshot is None:
        result.error = error or 'فشل الاتصال بالجهاز'
        device.connection_status = device.ConnectionStatus.ERROR
        device.last_error = result.error
        device.last_ping_at = timezone.now()
        device.save(update_fields=['connection_status', 'last_error', 'last_ping_at', 'updated_at'])
        return result

    result.serial_number = snapshot.serial_number
    result.firmware = snapshot.firmware
    if snapshot.serial_number:
        device.serial_number = snapshot.serial_number
    if snapshot.firmware:
        device.firmware_version = snapshot.firmware

    _persist_users(device, snapshot.users)
    name_map = _build_name_map(snapshot.users)
    enroll_map = _get_enrollment_map(device)

    result.users_on_device = len(snapshot.users)
    result.punches_fetched = len(snapshot.attendance)

    filtered = _filter_rows(snapshot.attendance, date_from, date_to)
    result.punches_after_filter = len(filtered)
    all_enriched = _enrich_punches(device, filtered, name_map, enroll_map)

    user_ids_in_punches = {p.device_user_id for p in all_enriched}
    result.unmapped_users = len(user_ids_in_punches - set(enroll_map.keys()))

    if import_db or dry_run:
        from apps.attendance.services.punch_sync import import_enriched_punches

        outcome = import_enriched_punches(
            device,
            all_enriched,
            dry_run=dry_run,
            incremental=use_incremental,
        )
        result.punches = (
            outcome['new_punches'] if use_incremental else all_enriched
        )
        result.punches_new = outcome['punches_new']
        result.skipped_time_filter = outcome.get('skipped_time_filter', 0)
        result.imported = outcome['imported']
        result.skipped_duplicate = outcome['skipped_duplicate']
        result.batch = outcome['batch']
    else:
        from apps.attendance.selectors.biometric_devices import make_sync_batch_label

        result.punches = all_enriched
        result.punches_new = len(all_enriched)
        result.batch = make_sync_batch_label(device)

    if export_path:
        result.export_path = str(export_punches_excel(result, export_path))

    result.ok = True
    return result
