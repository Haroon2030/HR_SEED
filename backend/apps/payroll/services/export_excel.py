"""تصدير مسير الرواتب إلى Excel — تنسيق كشف الرواتب (صف ترويسة + بيانات)."""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.utils import timezone

from dataclasses import dataclass

from apps.payroll.models import PayrollLine
from apps.payroll.services.payroll_line_columns import (
    HEADER_FILL_COLORS,
    MONEY_SUM_KEYS,
    PAYROLL_LINE_COLUMNS,
    build_ephemeral_payroll_line,
    lookup_source_payroll_line,
    payroll_lines_select_related,
    resolve_cell_value,
    resolve_detailed_allocation_cell_value,
)

PAYROLL_EXPORT_COLUMNS = PAYROLL_LINE_COLUMNS

COLUMN_WIDTHS = {
    'employee_number': 12,
    'employee_name': 22,
    'account_number': 24,
    'bank': 14,
    'account_type': 12,
    'salary_gross': 11,
    'id_number': 14,
    'branch': 12,
    'company': 16,
    'period_start': 12,
    'period_end': 12,
    'worked_days': 10,
    'basic_salary': 12,
    'earned_basic': 12,
    'housing_allowance': 11,
    'earned_housing': 13,
    'transport_allowance': 11,
    'fixed_other_allowance': 12,
    'additional': 10,
    'total_allowances': 12,
    'total_earnings': 14,
    'penalties_deductions': 14,
    'insurance_deduction': 16,
    'loan_deduction': 14,
    'total_deductions': 13,
    'net_salary': 11,
    'payment': 10,
}

ROW_HEIGHT = 18


def _money(val) -> float:
    if val is None or val == '':
        return 0.0
    if isinstance(val, Decimal):
        return float(val)
    return float(val)


def _write_payroll_sheet(ws, line_pairs, *, meta_note: str, resolve_value=None):
    """يكتب ترويسة وصفوف وإجماليات على ورقة واحدة. line_pairs: [(run, line), ...]."""
    get_cell_value = resolve_value or resolve_cell_value
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=False)
    header_font = Font(name='Arial', size=9, bold=True, color='000000')
    data_font = Font(name='Arial', size=9, color='000000')
    total_font = Font(name='Arial', size=9, bold=True, color='000000')

    header_row = 1
    data_start = 2

    for idx, (key, label, color_key, _col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        fill = PatternFill('solid', fgColor=HEADER_FILL_COLORS.get(color_key, 'B4C6E7'))
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(key, 11)

    ws.row_dimensions[header_row].height = 36
    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

    totals = {key: 0.0 for key in MONEY_SUM_KEYS}

    for i, (run, line) in enumerate(line_pairs):
        r = data_start + i
        for col_idx, (key, _label, _color, col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
            raw = get_cell_value(line, run, key)
            if col_type == 'text':
                val = raw if raw not in (None, '') else None
            elif col_type == 'days':
                val = float(raw) if raw not in (None, '') else None
            else:
                val = _money(raw) if raw not in (None, '') else None
                if val is not None:
                    totals[key] += val

            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if col_type == 'text':
                cell.alignment = align_right if key in ('employee_name', 'bank', 'company', 'branch') else align_center
            else:
                cell.alignment = align_center
            if col_type == 'money':
                cell.number_format = '#,##0.00'
            elif col_type == 'days':
                cell.number_format = '0.0'

        ws.row_dimensions[r].height = ROW_HEIGHT

    if line_pairs:
        footer_row = data_start + len(line_pairs)
        ws.row_dimensions[footer_row].height = ROW_HEIGHT
        for col_idx, (key, _label, _color, col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=footer_row, column=col_idx)
            cell.font = total_font
            cell.border = border
            cell.alignment = align_center
            if key == 'employee_name':
                cell.value = 'الإجمالي'
                cell.alignment = align_right
            elif col_type == 'money' and key in totals:
                cell.value = totals[key]
                cell.number_format = '#,##0.00'

    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = meta_note


def _payroll_line_pairs_for_runs(runs):
    if not runs:
        return []
    run_ids = [run.id for run in runs]
    lines = payroll_lines_select_related(
        PayrollLine.objects.filter(run_id__in=run_ids),
    ).select_related('run', 'run__branch', 'run__sponsorship').order_by(
        'run__branch__name', 'employee__name',
    )
    return [(line.run, line) for line in lines]


def build_payroll_run_workbook(run):
    """يُنشئ Workbook بتنسيق كشف الرواتب (ترويسة ملوّنة + صفوف الموظفين)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'كشف الرواتب'[:31]
    ws.sheet_view.rightToLeft = True

    line_pairs = [(run, line) for line in payroll_lines_select_related(run.lines).order_by('employee__name')]
    meta_note = (
        f'{run.branch.name if run.branch_id else ""} — {run.period_label} — '
        f'تصدير {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'
    )
    _write_payroll_sheet(ws, line_pairs, meta_note=meta_note)
    return wb


def build_payroll_runs_workbook(runs):
    """Workbook موحّد لعدة مسيرات (جدول واحد لكل الفروع المختارة)."""
    from openpyxl import Workbook

    runs = list(runs)
    if not runs:
        raise ValueError('لا توجد مسيرات للتصدير.')

    wb = Workbook()
    ws = wb.active
    ws.title = 'كشف الرواتب'[:31]
    ws.sheet_view.rightToLeft = True

    first = runs[0]
    branch_names = ', '.join(
        r.branch.name for r in runs if r.branch_id
    )[:120]
    meta_note = (
        f'{branch_names or "مسير موحّد"} — {first.period_label} — '
        f'تصدير {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'
    )
    _write_payroll_sheet(ws, _payroll_line_pairs_for_runs(runs), meta_note=meta_note)
    return wb


@dataclass
class DetailedPayrollExportRow:
    alloc_line: object
    payroll_line: object


def detailed_allocation_lines_for_run(run):
    return run.allocation_lines.select_related(
        'employee',
        'employee__branch',
        'employee__branch__company',
        'employee__bank',
        'employee__sponsorship',
        'branch',
        'from_branch',
    ).order_by(
        'employee__name',
        'bears_salary',
        'days_in_branch',
        'transfer_date',
        'id',
    )


def _payroll_line_for_allocation(alloc_line, run, *, cache: dict):
    emp_id = alloc_line.employee_id
    if emp_id not in cache:
        source = lookup_source_payroll_line(alloc_line, run)
        cache[emp_id] = source or build_ephemeral_payroll_line(alloc_line.employee, run)
    return cache[emp_id]


def detailed_payroll_export_pairs(run):
    """صف لكل سطر توزيع فرع — بأعمدة كشف الرواتب."""
    cache = {}
    pairs = []
    for alloc_line in detailed_allocation_lines_for_run(run):
        payroll_line = _payroll_line_for_allocation(alloc_line, run, cache=cache)
        export_row = DetailedPayrollExportRow(
            alloc_line=alloc_line,
            payroll_line=payroll_line,
        )
        pairs.append((run, export_row))
    return pairs


def _resolve_detailed_export_row(export_row: DetailedPayrollExportRow, run, key: str):
    return resolve_detailed_allocation_cell_value(
        export_row.alloc_line,
        run,
        key,
        export_row.payroll_line,
    )


def build_payroll_detailed_run_workbook(run):
    """Workbook تفصيلي بأعمدة كشف الرواتب — صف لكل فرع في توزيع النقل."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'كشف الرواتب'[:31]
    ws.sheet_view.rightToLeft = True

    company_name = run.company.name if run.company_id else 'شركة'
    meta_note = (
        f'{company_name} — {run.period_label} — مسير تفصيلي — '
        f'تصدير {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'
    )
    _write_payroll_sheet(
        ws,
        detailed_payroll_export_pairs(run),
        meta_note=meta_note,
        resolve_value=_resolve_detailed_export_row,
    )
    return wb


def payroll_run_excel_filename(run) -> str:
    branch = run.branch_id or 'run'
    return f'payroll_{branch}_{run.period_year}_{run.period_month:02d}.xlsx'


def payroll_detailed_run_excel_filename(run) -> str:
    company = (run.company.name if run.company_id else 'detailed').replace(' ', '_')[:40]
    return f'payroll_detailed_{company}_{run.period_year}_{run.period_month:02d}.xlsx'


def payroll_runs_excel_filename(*, year: int, month: int, salary_mode: str) -> str:
    return f'payroll_{year}_{month:02d}_{salary_mode}.xlsx'


def workbook_to_response(wb, filename: str):
    from django.http import HttpResponse

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
