"""استيراد الأرصدة الافتتاحية من ملف Excel بعد الترحيل من نظام سابق."""
from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.employees.models import Employee
from apps.employees.services.migration_balance import global_cutover_date, parse_cutover_date
from apps.employees.services.opening_balances import apply_opening_balance_to_employee

# عناوين الأعمدة المقبولة (عربي / إنجليزي)
COL_EMPLOYEE = frozenset({
    'employee_number', 'رقم_الموظف', 'رقم الموظف', 'employee_id', 'id',
})
COL_LEAVE = frozenset({
    'opening_leave_days', 'رصيد_إجازة', 'رصيد_إجازة_افتتاحي', 'رصيد الاجازة', 'leave_days',
})
COL_EOSB = frozenset({
    'opening_eosb_amount', 'مخصص_eosb', 'مخصص_نهاية_الخدمة', 'eosb', 'eosb_amount',
})
COL_ACCRUAL_START = frozenset({
    'leave_accrual_start_date', 'تاريخ_بدء_الإجازة', 'تاريخ الانتقال', 'cutover_date',
})


def _normalize_header(value) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = re.sub(r'\s+', '_', text)
    return text


def _map_headers(row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = _normalize_header(cell)
        if key in COL_EMPLOYEE:
            mapping['employee'] = idx
        elif key in COL_LEAVE:
            mapping['leave'] = idx
        elif key in COL_EOSB:
            mapping['eosb'] = idx
        elif key in COL_ACCRUAL_START:
            mapping['accrual_start'] = idx
    return mapping


def _cell_decimal(row, idx: int | None, default=Decimal('0')) -> Decimal:
    if idx is None or idx >= len(row):
        return default
    raw = row[idx]
    if raw is None or str(raw).strip() == '':
        return default
    try:
        return Decimal(str(raw).replace(',', '').strip()).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        raise ValueError(f'قيمة رقمية غير صالحة: {raw!r}')


def _cell_date(row, idx: int | None) -> date | None:
    if idx is None or idx >= len(row):
        return None
    raw = row[idx]
    if raw is None or str(raw).strip() == '':
        return None
    if isinstance(raw, date):
        return raw
    from datetime import datetime
    if isinstance(raw, datetime):
        return raw.date()
    text = str(raw).strip()[:10]
    return date.fromisoformat(text)


def _resolve_employee(identifier: str) -> Employee | None:
    ident = (identifier or '').strip()
    if not ident:
        return None
    if ident.isdigit():
        emp = Employee.objects.filter(pk=int(ident), is_deleted=False).first()
        if emp:
            return emp
    return Employee.objects.filter(
        employee_number=ident,
        is_deleted=False,
    ).first() or Employee.objects.filter(
        id_number=ident,
        is_deleted=False,
    ).first()


class Command(BaseCommand):
    help = (
        'استيراد رصيد إجازة افتتاحي ومخصص EOSB من Excel. '
        'الأعمدة: رقم_الموظف، رصيد_إجازة_افتتاحي، مخصص_eosb (اختياري: تاريخ_بدء_الإجازة).'
    )

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='مسار ملف Excel (.xlsx)')
        parser.add_argument(
            '--cutover-date',
            type=str,
            default='',
            help='تاريخ الانتقال YYYY-MM-DD (يفضّل على HR_MIGRATION_CUTOVER_DATE)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='معاينة فقط بدون حفظ',
        )
        parser.add_argument(
            '--replace',
            action='store_true',
            help='استبدال أرصدة موظفين معتمدين مسبقاً',
        )

    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'الملف غير موجود: {path}')

        cutover = parse_cutover_date(options.get('cutover_date')) or global_cutover_date()
        if not cutover:
            raise CommandError(
                'حدّد --cutover-date أو عيّن HR_MIGRATION_CUTOVER_DATE في البيئة.'
            )

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('مكتبة openpyxl غير مثبتة.') from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('الملف فارغ.')

        headers = _map_headers(rows[0])
        if 'employee' not in headers:
            raise CommandError('عمود رقم الموظف مفقود (employee_number / رقم_الموظف).')
        if 'leave' not in headers and 'eosb' not in headers:
            raise CommandError('يجب وجود عمود رصيد إجازة أو مخصص EOSB على الأقل.')

        ok = 0
        errors: list[str] = []

        for line_no, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            ident = str(row[headers['employee']] or '').strip()
            try:
                emp = _resolve_employee(ident)
                if not emp:
                    errors.append(f'سطر {line_no}: موظف غير موجود ({ident})')
                    continue

                leave_days = _cell_decimal(row, headers.get('leave'))
                eosb_amount = _cell_decimal(row, headers.get('eosb'))
                accrual_start = _cell_date(row, headers.get('accrual_start'))

                if options['dry_run']:
                    self.stdout.write(
                        f'[معاينة] {emp.name}: إجازة={leave_days} EOSB={eosb_amount} '
                        f'انتقال={accrual_start or cutover}'
                    )
                    ok += 1
                    continue

                with transaction.atomic():
                    apply_opening_balance_to_employee(
                        emp,
                        opening_leave_days=leave_days,
                        opening_eosb_amount=eosb_amount,
                        cutover_date=cutover,
                        leave_accrual_start_date=accrual_start,
                        replace_existing=options['replace'],
                    )
                ok += 1
                self.stdout.write(self.style.SUCCESS(f'سطر {line_no}: {emp.name} — تم'))
            except Exception as exc:
                errors.append(f'سطر {line_no} ({ident}): {exc}')

        wb.close()

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'نجح: {ok}'))
        if errors:
            self.stdout.write(self.style.ERROR(f'أخطاء: {len(errors)}'))
            for err in errors:
                self.stdout.write(self.style.ERROR(f'  • {err}'))
            raise CommandError('اكتمل الاستيراد مع أخطاء.')

        self.stdout.write(self.style.SUCCESS('اكتمل الاستيراد بنجاح.'))
